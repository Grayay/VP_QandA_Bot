from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


REQUIRED_COLUMNS = ("Вопрос", "Ответ", "Отдел", "Раздел")
EMPTY_SECTION_TITLE = "Без раздела"
TRANSFER_TO_BOOKER_ANSWER = "Перевести на букера"


class FAQLoaderError(RuntimeError):
    pass


@dataclass(frozen=True)
class FAQItem:
    id: int
    token: str
    question: str
    answer: str
    department: str | None
    section: str | None


@dataclass(frozen=True)
class FAQData:
    items: list[FAQItem]
    sections: list[str]
    by_section: dict[str, list[FAQItem]]
    by_id: dict[int, FAQItem]
    skipped_rows: int = 0


def load_faq(path: Path) -> FAQData:
    if not path.exists():
        raise FAQLoaderError(f"Файл {path.name} не найден в корне проекта.")

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError) as error:
        raise FAQLoaderError(f"Не удалось открыть Excel-файл: {error}") from error

    try:
        sheet = workbook.active

        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            raise FAQLoaderError("В Excel-файле нет строки с заголовками.")

        headers = _trim_empty_tail([_normalize_cell(value) for value in header_row])
        _validate_headers(headers)

        column_index = {name: headers.index(name) for name in REQUIRED_COLUMNS}
        items: list[FAQItem] = []
        sections: list[str] = []
        by_section: dict[str, list[FAQItem]] = {}
        skipped_rows = 0

        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            question = _get_value(row, column_index["Вопрос"])
            answer = _get_value(row, column_index["Ответ"])
            department = _get_value(row, column_index["Отдел"])
            section = _get_value(row, column_index["Раздел"])

            if not question and not answer and not department and not section:
                skipped_rows += 1
                continue

            if not question or not answer:
                raise FAQLoaderError(
                    f"В строке {row_number} должны быть заполнены Вопрос и Ответ."
                )

            item = FAQItem(
                id=len(items) + 1,
                token=str(len(items) + 1),
                question=question,
                answer=answer,
                department=department or None,
                section=section or None,
            )
            items.append(item)

            section_title = section_title_for_value(section)
            if section_title not in by_section:
                by_section[section_title] = []
                sections.append(section_title)
            by_section[section_title].append(item)

        if not items:
            raise FAQLoaderError("В Excel-файле нет FAQ-вопросов.")

        return FAQData(
            items=items,
            sections=sections,
            by_section=by_section,
            by_id={item.id: item for item in items},
            skipped_rows=skipped_rows,
        )
    finally:
        workbook.close()


def _validate_headers(headers: list[str]) -> None:
    expected = set(REQUIRED_COLUMNS)
    actual = set(headers)

    missing = expected - actual

    if missing:
        details = []
        details.append("нет колонок: " + ", ".join(sorted(missing)))
        raise FAQLoaderError("Неверные колонки в Excel-файле: " + "; ".join(details))


def _get_value(row: tuple[object, ...], index: int) -> str:
    if index >= len(row):
        return ""
    return _normalize_cell(row[index])


def _normalize_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _trim_empty_tail(values: list[str]) -> list[str]:
    while values and not values[-1]:
        values.pop()
    return values


def section_title_for_value(section: str | None) -> str:
    value = (section or "").strip()
    return value or EMPTY_SECTION_TITLE


def is_transfer_to_booker_answer(answer: str) -> bool:
    normalized = answer.strip().lower().replace("ё", "е")
    expected = TRANSFER_TO_BOOKER_ANSWER.lower().replace("ё", "е")
    return normalized == expected
